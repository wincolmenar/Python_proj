import os
from openpyxl import Workbook

def extract_and_trim(input_folder):
    for filename in os.listdir(input_folder):
        if filename.endswith(".txt"):
            # Read content from the TXT file
            with open(os.path.join(input_folder, filename), 'r') as txt_file:
                content = txt_file.read().splitlines()

            # Create a new Excel workbook
            wb = Workbook()
            ws = wb.active
            mm_yy = ["01.21","02.21","03.21","04.21","05.21","06.21","07.21","08.21","09.21","10.21","11.21","12.21","01.22","01.23","01.24","02.22","02.23","02.24","03.22","03.23","03.24","04.22","04.23","04.24","05.22","05.23","05.24","06.22","06.23","06.24","07.22","07.23","07.24","08.22","08.23","08.24","09.22","09.23","09.24","10.22","10.23","10.24","11.22","11.23","11.24","12.22","12.23","12.24",] ### increment of 3 years in mm.yy format
            

            # Apply trimming logic
            for row_num, cell_value in enumerate(content, start=1):
                    if any(keyword in cell_value for keyword in ["General Ledger from the Document File", "RFHABU00"]):
                        ws[f"O{row_num}"] = cell_value[:40]
                        ws[f"T{row_num}"] = cell_value[40:96]
                        ws[f"Z{row_num}"] = cell_value[98:116]
                        ws[f"AA{row_num}"] = cell_value[116:]

                    elif "-------------------------------------------------------------------------------------" in cell_value:
                        # ws[f"O{row_num}"] = cell_value[:38]
                        ws[f"O{row_num}"] = cell_value

                    elif cell_value.startswith("0087") or cell_value.startswith("1508") or cell_value.startswith("1870") and "Business area totals" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:4]
                        ws[f"P{row_num}"] = cell_value[7:20]
                        ws[f"Q{row_num}"] = cell_value[21:24]
                        ws[f"R{row_num}"] = cell_value[27:]

                    elif cell_value.startswith("0087") or cell_value.startswith("1508") or cell_value.startswith("1870") and "Account totals" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:4]
                        ws[f"P{row_num}"] = cell_value[8:15]
                        ws[f"Q{row_num}"] =cell_value[16:]
                        # ws[f"R{row_num}"] = cell_value[22:]
                    
                    elif cell_value.startswith("0087") or cell_value.startswith("1508") or cell_value.startswith("1870") and "Company code totals" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:4]
                        ws[f"P{row_num}"] = cell_value[5:11]
                        ws[f"Q{row_num}"] = cell_value[11:30]

                    elif cell_value.startswith("0087") or cell_value.startswith("1508") or cell_value.startswith("1870") and "PHP" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:4]
                        ws[f"P{row_num}"] = cell_value[7:21]
                        ws[f"Q{row_num}"] = cell_value[21:24]
                        ws[f"R{row_num}"] = cell_value[27:69]

                    elif cell_value.startswith("0087") or cell_value.startswith("1508") or cell_value.startswith("1870") and "USD" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:4]
                        ws[f"P{row_num}"] = cell_value[7:21]
                        ws[f"Q{row_num}"] = cell_value[21:24]
                        ws[f"R{row_num}"] = cell_value[27:69]

                    elif cell_value.startswith("0087") or cell_value.startswith("1508") or cell_value.startswith("1870") and "EUR" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:4]
                        ws[f"P{row_num}"] = cell_value[7:21]
                        ws[f"Q{row_num}"] = cell_value[21:24]
                        ws[f"R{row_num}"] = cell_value[27:69]

                    elif "Totals accross all company codes" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:6]
                        ws[f"P{row_num}"] = cell_value[7:38]

                    elif "Number of master records read:" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:30]
                        ws[f"P{row_num}"] = cell_value[30:]

                    elif "Number of items read:" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:21]
                        ws[f"P{row_num}"] = cell_value[21:74]

                    elif "                        Cost center" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:37]
                        ws[f"P{row_num}"] = cell_value[37:48]
                        ws[f"Q{row_num}"] = cell_value[48:52]

                    elif "                        Order Number" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:40]
                        ws[f"P{row_num}"] = cell_value[40:52]

                    elif "                        Profit center" in cell_value:
                        ws[f"P{row_num}"] = cell_value[0:39]
                        ws[f"Q{row_num}"] = cell_value[39:49]

                    elif "                        Personnel numbe" in cell_value:
                        ws[f"P{row_num}"] = cell_value[0:39] + "r:"
                        ws[f"Q{row_num}"] = cell_value[40:48]

                    elif "                        Purchase order" in cell_value:
                        ws[f"P{row_num}"] = cell_value[0:39]
                        ws[f"Q{row_num}"] = cell_value[39:50]
                        ws[f"R{row_num}"] = cell_value[50:66] + "tem no."
                        ws[f"S{row_num}"] = cell_value[66:71]

                    elif "                        Quant" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:29]
                        ws[f"P{row_num}"] = cell_value[29:47]
                        ws[f"Q{row_num}"] = cell_value[47:61]
                        ws[f"R{row_num}"] = cell_value[61:66]

                    
                    elif any(element in cell_value and "PHP" in cell_value for element in mm_yy):
                        ws[f"O{row_num}"] = cell_value[0:6]
                        ws[f"P{row_num}"] = cell_value[6:13]
                        ws[f"Q{row_num}"] = cell_value[13:24]
                        ws[f"R{row_num}"] = cell_value[24:27]
                        ws[f"S{row_num}"] = cell_value[27:30]
                        ws[f"T{row_num}"] = cell_value[30:49]
                        ws[f"U{row_num}"] = cell_value[49:56]
                        ws[f"V{row_num}"] = cell_value[56:73]
                        ws[f"W{row_num}"] = cell_value[73:78]
                        ws[f"X{row_num}"] = cell_value[78:80]
                        ws[f"Y{row_num}"] = cell_value[80:91]
                        ws[f"Z{row_num}"] = cell_value[91:102]
                        ws[f"AA{row_num}"] = cell_value[102:122]

                    elif "Totals                             Debit             Credit" in cell_value or "Balance Carryforward" in cell_value or "New balance" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:35]
                        ws[f"P{row_num}"] = cell_value[35:53]
                        ws[f"Q{row_num}"] = cell_value[53:72]
                        ws[f"R{row_num}"] = cell_value[72:80]
                        ws[f"S{row_num}"] = cell_value[80:98]

                    elif "Pstng Pstng  Document" in cell_value or "per.  date   number" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:6]
                        ws[f"P{row_num}"] = cell_value[6:13]
                        ws[f"Q{row_num}"] = cell_value[13:24]
                        ws[f"R{row_num}"] = cell_value[24:27]
                        ws[f"S{row_num}"] = cell_value[27:30]
                        ws[f"T{row_num}"] = cell_value[30:49]
                        ws[f"U{row_num}"] = cell_value[49:56]
                        ws[f"V{row_num}"] = cell_value[56:73]
                        ws[f"W{row_num}"] = cell_value[73:78]
                        ws[f"X{row_num}"] = cell_value[78:80]
                        ws[f"Y{row_num}"] = cell_value[80:91]
                        ws[f"Z{row_num}"] = cell_value[91:102]
                        ws[f"AA{row_num}"] = cell_value[102:122]


                    elif any(element in cell_value for element in mm_yy):
                        ws[f"O{row_num}"] = cell_value[0:13]
                        ws[f"P{row_num}"] = cell_value[13:39]
                        ws[f"Q{row_num}"] = cell_value[39:58]
                        ws[f"R{row_num}"] = cell_value[58:78]
                        ws[f"S{row_num}"] = cell_value[78:97]

                    elif 3 <= len(cell_value) and len(cell_value) <= 15:
                        ws[f"O{row_num}"] = cell_value

                    else:
                        ws[f"P{row_num}"] = cell_value

            # Delete columns A to N
            ws.delete_cols(1, 14)

            # Apply trim function to all cells
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        cell.value = cell.value.strip()


            # Save the Excel workbook with the same name as the original TXT file
            output_filename = os.path.splitext(filename)[0] + "_trimmed.xlsx"
            output_path = os.path.join(input_folder, output_filename)
            wb.save(output_path)
            print(f"Conversion and trimming successful: {filename} -> {output_filename}")

# Specify the path to the folder containing TXT files
input_folder_path = r"C:\Users\gnjwv\OneDrive - Bayer\Desktop\General Ledger" ### change to dynamic
extract_and_trim(input_folder_path)