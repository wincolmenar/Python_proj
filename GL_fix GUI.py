import os
from tkinter import *
from openpyxl import load_workbook

# folder path of the input and output
input_path = os.path.join(os.path.expanduser("~"), "OneDrive - Bayer", "Desktop", "Convert GL", "Input Folder")
output_path = os.path.join(os.path.expanduser("~"), "OneDrive - Bayer", "Desktop", "Convert GL", "Output Folder")

# Creates a  pop-up window
root = Tk()
root.title("GL Converter")

status = Label(root, text = "Ready")
status.pack(side = BOTTOM, fill = X)

# method for the status
def update_status(message):
    status.config(text = message)

def fix_general_journal():

    # Iterate through each file in the folder
    for filename in os.listdir(input_path):
        try:
            if filename.endswith(".xlsx"):
    
                # Load the workbook, script will scan all files inside the folder regardless of the name (include as long as it's .xlsx type)
                wb = load_workbook(os.path.join(input_path, filename))
    
                # Select the worksheet "General_Journal_Raw"
                ws_raw = wb["General_Journal_Raw"]
    
                # Create a new worksheet for the results
                ws_results = wb.create_sheet("Trimmed_Results")
    
                # Iterate through each cell in column A of the "General_Journal_Raw" sheet
                for cell in ws_raw["A"]:
                    cell_value = cell.value

                    if cell_value is not None:
                        # VBA trimming logic
                        if "Document Journal" in cell_value:
                            ws_results[f"O{cell.row}"] = cell_value[:38]
                            ws_results[f"T{cell.row}"] = cell_value[38:81]
                            ws_results[f"Z{cell.row}"] = cell_value[82:112]
                            ws_results[f"AA{cell.row}"] = cell_value[113:134]
                        
                        elif "RFBELJ10" in cell_value:
                            ws_results[f"O{cell.row}"] = cell_value[:96]
                            ws_results[f"Z{cell.row}"] = cell_value[96:116]
                            ws_results[f"AA{cell.row}"] = cell_value[116:]
                        
                        elif "-------------------------------------------------------------------------------------" in cell_value:
                            ws_results[f"O{cell.row}"] = cell_value

                        elif any(keyword in cell_value for keyword in ["Carryfwd", "Pages Total", "Cumulated"]):
                            ws_results[f"Y{cell.row}"] = cell_value[69:83]
                            ws_results[f"Z{cell.row}"] = cell_value[98:114]
                            ws_results[f"AA{cell.row}"] = cell_value[114:130]

                        elif "Account Name.............." in cell_value:
                            ws_results[f"P{cell.row}"] = cell_value[9:21]
                            ws_results[f"Q{cell.row}"] = cell_value[35:38]
                            ws_results[f"R{cell.row}"] = cell_value[42:43]
                            ws_results[f"S{cell.row}"] = cell_value[44:54]
                            ws_results[f"T{cell.row}"] = cell_value[55:57]
                            ws_results[f"U{cell.row}"] = cell_value[65:75]
                            ws_results[f"V{cell.row}"] = cell_value[60:62]
                            ws_results[f"W{cell.row}"] = cell_value[76:78]
                            ws_results[f"X{cell.row}"] = cell_value[82:95]
                            ws_results[f"Y{cell.row}"] = cell_value[95:99]
                            ws_results[f"Z{cell.row}"] = cell_value[99:115]
                            ws_results[f"AA{cell.row}"] = cell_value[115:130]

                        elif any(keyword in cell_value for keyword in ["Year Curr", "PHP   0087", "PHP   **** ** ", "                   T                                   T "]):
                            ws_results[f"O{cell.row}"] = cell_value[:4]
                            ws_results[f"P{cell.row}"] = cell_value[5:11]
                            ws_results[f"Q{cell.row}"] = cell_value[11:15]
                            ws_results[f"R{cell.row}"] = cell_value[16:18]
                            ws_results[f"S{cell.row}"] = cell_value[19:21]
                            ws_results[f"T{cell.row}"] = cell_value[21:37]
                            ws_results[f"U{cell.row}"] = cell_value[38:54]
                            ws_results[f"V{cell.row}"] = cell_value[55:57]
                            ws_results[f"W{cell.row}"] = cell_value[57:73]
                            ws_results[f"X{cell.row}"] = cell_value[73:91]
                            ws_results[f"Y{cell.row}"] = cell_value[90:92]
                            ws_results[f"Z{cell.row}"] = cell_value[92:109]
                            ws_results[f"AA{cell.row}"] = cell_value[109:127]

                        elif "Seq.no.  CPU" in cell_value:
                            ws_results[f"O{cell.row}"] = cell_value[:8]
                            ws_results[f"P{cell.row}"] = cell_value[9:15]
                            ws_results[f"Q{cell.row}"] = cell_value[16:26]
                            ws_results[f"R{cell.row}"] = cell_value[27:33]
                            ws_results[f"S{cell.row}"] = cell_value[34:40]
                            ws_results[f"T{cell.row}"] = cell_value[41:61]
                            ws_results[f"U{cell.row}"] = cell_value[61:81]

                        elif cell_value[:4] == "0000":
                            ws_results[f"O{cell.row}"] = cell_value[:8]
                            ws_results[f"P{cell.row}"] = cell_value[9:15]
                            ws_results[f"Q{cell.row}"] = cell_value[16:26]
                            ws_results[f"R{cell.row}"] = cell_value[27:33]
                            ws_results[f"S{cell.row}"] = cell_value[34:40]
                            ws_results[f"T{cell.row}"] = cell_value[41:61]
                            ws_results[f"U{cell.row}"] = cell_value[61:101]

                        else:
                            ws_results[f"P{cell.row}"] = cell_value[9:35]
                            ws_results[f"Q{cell.row}"] = cell_value[35:38]
                            ws_results[f"R{cell.row}"] = cell_value[42:43]
                            ws_results[f"S{cell.row}"] = cell_value[44:54]
                            ws_results[f"T{cell.row}"] = cell_value[55:57]
                            ws_results[f"U{cell.row}"] = cell_value[65:75]
                            ws_results[f"V{cell.row}"] = cell_value[60:62]
                            ws_results[f"W{cell.row}"] = cell_value[76:78]
                            ws_results[f"X{cell.row}"] = cell_value[79:95]
                            ws_results[f"Y{cell.row}"] = cell_value[95:99]
                            ws_results[f"Z{cell.row}"] = cell_value[99:115]
                            ws_results[f"AA{cell.row}"] = cell_value[115:130]

                # Delete Columns A to N
                ws_results.delete_cols(1, 14)

                # Trim every cells in a row
                for row in ws_results.iter_rows():
                    for cell in row:
                        if cell.value is not None and isinstance(cell.value, str):
                            cell.value = cell.value.strip()
            # Save the workbook with the new data
            wb.save(os.path.join(output_path, "Converted " + filename))
            update_status("Done. You can close this window.")

        except Exception as e:
            print(f"Error {e}")

# for button in closing the pop window
def close_window():
    root.destroy()

# content of the window
Button1 = Button(root, text = "Convert", command = fix_general_journal, padx = 30)
# shoving it into the window
Button1.pack(side = "left")

# content of the window
Button2 = Button(root, text = "Close", command = close_window, padx = 30)
# shoving it into the window
Button2.pack(side = "right")

# Pop the window open until something is done
root.mainloop()