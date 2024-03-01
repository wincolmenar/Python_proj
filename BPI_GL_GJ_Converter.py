import os
from tkinter import *
from datetime import datetime
from openpyxl import Workbook

# Specify the path of the GL files, in this case the Input and Output folder in Conver GL folder
GL_input_path_folder = os.path.join(os.path.expanduser("~"), "OneDrive - Bayer", "Desktop", "File Converter", "GL Input Folder")
GL_output_path_folder = os.path.join(os.path.expanduser("~"), "OneDrive - Bayer", "Desktop", "File Converter", "GL Output Folder")

# Specify the path of the GJ files, in this case the Input and Output folder in Conver GL folder
GJ_input_folder_path = os.path.join(os.path.expanduser("~"), "OneDrive - Bayer", "Desktop", "File Converter", "GJ Input Folder")
GJ_output_folder_path = os.path.join(os.path.expanduser("~"), "OneDrive - Bayer", "Desktop", "File Converter", "GJ Output Folder")


## GUI script
# Create the window
root = Tk()
root.title("GL GJ Converter")

status_label = Label(root, text="Ready")
# fill = x makes it center
status_label.pack(side = BOTTOM, fill = X, pady = 5)


# function for the status
def update_status(message):
    status_label.config(text = message)


# generate months for the mm.yy
def generate_yymm(start_year, end_year):
    # empty list to store the mm.yy values
    mm_yy_series = []

    # year loop.  Python generates sequence of numbers up to, but not including, the stop. True number is until 2040
    for year in range(start_year, end_year):
        # month loop. Python generates sequence of numbers up to, but not including, the stop. True number is until 12
        for month in range(1, 13):
            indiv_date_start = datetime(year, month, 1)
            mm_yy_series.append(indiv_date_start.strftime('%m.%y'))
    
    # return and show the list of mm.yy values
    return mm_yy_series


# Main GL trim function
def GL_extract_and_trim(input_folder):
    for filename in os.listdir(input_folder):
        try:
            if filename.endswith(".txt"):
                # Read content from the TXT file
                with open(os.path.join(input_folder, filename), 'r') as txt_file:
                    content = txt_file.read().splitlines()

                # Create a new Excel workbook
                wb = Workbook()
                ws = wb.active
                mm_yy = generate_yymm(2021, 2041)
                
                # Apply trimming logic
                for row_num, cell_value in enumerate(content, start=1):
                    if any(keyword in cell_value for keyword in ["General Ledger from the Document File", "RFHABU00"]):
                        ws[f"O{row_num}"] = cell_value[:40]
                        ws[f"T{row_num}"] = cell_value[40:96]
                        ws[f"Z{row_num}"] = cell_value[98:116]
                        ws[f"AA{row_num}"] = cell_value[116:]

                    elif "-------------------------------------------------------------------------------------" in cell_value:
                        # ws[f"O{row_num}"] = cell_value[:38]
                        ws[f"O{row_num}"] = cell_value

                    elif cell_value.startswith("0087") or cell_value.startswith("1508") or cell_value.startswith("1870") and "Business area totals" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:4]
                        ws[f"P{row_num}"] = cell_value[7:20]
                        ws[f"Q{row_num}"] = cell_value[21:24]
                        ws[f"R{row_num}"] = cell_value[27:]

                    elif cell_value.startswith("0087") or cell_value.startswith("1508") or cell_value.startswith("1870") and "Account totals" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:4]
                        ws[f"P{row_num}"] = cell_value[8:15]
                        ws[f"Q{row_num}"] =cell_value[16:19]
                        ws[f"R{row_num}"] = cell_value[22:]
                    
                    elif cell_value.startswith("0087") or cell_value.startswith("1508") or cell_value.startswith("1870") and "Company code totals" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:4]
                        ws[f"P{row_num}"] = cell_value[5:11]
                        ws[f"Q{row_num}"] = cell_value[11:30]

                    elif cell_value.startswith("0087") or cell_value.startswith("1508") or cell_value.startswith("1870") and "PHP" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:4]
                        ws[f"P{row_num}"] = cell_value[7:21]
                        ws[f"Q{row_num}"] = cell_value[21:24]
                        ws[f"R{row_num}"] = cell_value[27:69]

                    elif cell_value.startswith("0087") or cell_value.startswith("1508") or cell_value.startswith("1870") and "USD" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:4]
                        ws[f"P{row_num}"] = cell_value[7:21]
                        ws[f"Q{row_num}"] = cell_value[21:24]
                        ws[f"R{row_num}"] = cell_value[27:69]

                    elif cell_value.startswith("0087") or cell_value.startswith("1508") or cell_value.startswith("1870") and "EUR" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:4]
                        ws[f"P{row_num}"] = cell_value[7:21]
                        ws[f"Q{row_num}"] = cell_value[21:24]
                        ws[f"R{row_num}"] = cell_value[27:69]

                    elif "Totals accross all company codes" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:6]
                        ws[f"P{row_num}"] = cell_value[7:38]

                    elif "Number of master records read:" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:30]
                        ws[f"P{row_num}"] = cell_value[30:]

                    elif "Number of items read:" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:21]
                        ws[f"P{row_num}"] = cell_value[21:74]

                    elif "                        Cost center" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:37]
                        ws[f"P{row_num}"] = cell_value[37:48]
                        ws[f"Q{row_num}"] = cell_value[48:52]

                    elif "                        Order Number" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:40]
                        ws[f"P{row_num}"] = cell_value[40:52]

                    elif "                        Profit center" in cell_value:
                        ws[f"P{row_num}"] = cell_value[0:39]
                        ws[f"Q{row_num}"] = cell_value[39:49]

                    elif "                        Personnel numbe" in cell_value:
                        ws[f"P{row_num}"] = cell_value[0:39] + "r:"
                        ws[f"Q{row_num}"] = cell_value[40:48]

                    elif "                        Purchase order" in cell_value:
                        ws[f"P{row_num}"] = cell_value[0:39]
                        ws[f"Q{row_num}"] = cell_value[39:50]
                        ws[f"R{row_num}"] = cell_value[50:66] + "tem no."
                        ws[f"S{row_num}"] = cell_value[66:71]

                    elif "                        Quant" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:29]
                        ws[f"P{row_num}"] = cell_value[29:47]
                        ws[f"Q{row_num}"] = cell_value[47:61]
                        ws[f"R{row_num}"] = cell_value[61:66]

                    
                    elif any(element in cell_value and "PHP" in cell_value for element in mm_yy):
                        ws[f"O{row_num}"] = cell_value[0:6]
                        ws[f"P{row_num}"] = cell_value[6:13]
                        ws[f"Q{row_num}"] = cell_value[13:24]
                        ws[f"R{row_num}"] = cell_value[24:27]
                        ws[f"S{row_num}"] = cell_value[27:30]
                        ws[f"T{row_num}"] = cell_value[30:49]
                        ws[f"U{row_num}"] = cell_value[49:56]
                        ws[f"V{row_num}"] = cell_value[56:73]
                        ws[f"W{row_num}"] = cell_value[73:78]
                        ws[f"X{row_num}"] = cell_value[78:80]
                        ws[f"Y{row_num}"] = cell_value[80:91]
                        ws[f"Z{row_num}"] = cell_value[91:102]
                        ws[f"AA{row_num}"] = cell_value[102:122]

                    elif "Totals                             Debit             Credit" in cell_value or "Balance Carryforward" in cell_value or "New balance" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:35]
                        ws[f"P{row_num}"] = cell_value[35:53]
                        ws[f"Q{row_num}"] = cell_value[53:72]
                        ws[f"R{row_num}"] = cell_value[72:80]
                        ws[f"S{row_num}"] = cell_value[80:98]

                    elif "Pstng Pstng  Document" in cell_value or "per.  date   number" in cell_value:
                        ws[f"O{row_num}"] = cell_value[0:6]
                        ws[f"P{row_num}"] = cell_value[6:13]
                        ws[f"Q{row_num}"] = cell_value[13:24]
                        ws[f"R{row_num}"] = cell_value[24:27]
                        ws[f"S{row_num}"] = cell_value[27:30]
                        ws[f"T{row_num}"] = cell_value[30:49]
                        ws[f"U{row_num}"] = cell_value[49:56]
                        ws[f"V{row_num}"] = cell_value[56:73]
                        ws[f"W{row_num}"] = cell_value[73:78]
                        ws[f"X{row_num}"] = cell_value[78:80]
                        ws[f"Y{row_num}"] = cell_value[80:91]
                        ws[f"Z{row_num}"] = cell_value[91:102]
                        ws[f"AA{row_num}"] = cell_value[102:122]


                    elif any(element in cell_value for element in mm_yy):
                        ws[f"O{row_num}"] = cell_value[0:13]
                        ws[f"P{row_num}"] = cell_value[13:39]
                        ws[f"Q{row_num}"] = cell_value[39:58]
                        ws[f"R{row_num}"] = cell_value[58:78]
                        ws[f"S{row_num}"] = cell_value[78:97]

                    elif 3 <= len(cell_value) and len(cell_value) <= 15:
                        ws[f"O{row_num}"] = cell_value

                    else:
                        ws[f"P{row_num}"] = cell_value

                # Delete columns A to N
                ws.delete_cols(1, 14)

                # Apply trim function to all cells
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None and isinstance(cell.value, str):
                            cell.value = cell.value.strip()


                # Save the Excel workbook with the same name as the original TXT file
                output_filename = os.path.splitext(filename)[0] + "_GL trimmed.xlsx"
                output_path = os.path.join(GL_output_path_folder, output_filename)
                wb.save(output_path)

        except Exception as e:
            update_status(f"Error {e}.")


## Main GJ trim function
def GJ_extract_and_trim(input_folder):
    for filename in os.listdir(input_folder):
        try:
            if filename.endswith(".txt"):
                # Read content from the TXT file
                with open(os.path.join(input_folder, filename), 'r') as txt_file:
                    content = txt_file.read().splitlines()
    
                # Create a new Excel workbook
                wb = Workbook()
                ws = wb.active
    
                # Apply trimming logic
                for row_num, cell_value in enumerate(content, start=1):
                        if "Document Journal" in cell_value:
                            ws[f"O{row_num}"] = cell_value[:38]
                            ws[f"T{row_num}"] = cell_value[38:81]
                            ws[f"Z{row_num}"] = cell_value[82:112]
                            ws[f"AA{row_num}"] = cell_value[113:134]
                        elif "RFBELJ10" in cell_value:
                            ws[f"O{row_num}"] = cell_value[:96]
                            ws[f"Z{row_num}"] = cell_value[96:116]
                            ws[f"AA{row_num}"] = cell_value[116:]
                        elif "-------------------------------------------------------------------------------------" in cell_value:
                            ws[f"O{row_num}"] = cell_value
                        elif any(keyword in cell_value for keyword in ["Carryfwd", "Pages Total", "Cumulated"]):
                            ws[f"Y{row_num}"] = cell_value[69:83]
                            ws[f"Z{row_num}"] = cell_value[98:114]
                            ws[f"AA{row_num}"] = cell_value[114:130]
                        elif any(keyword in cell_value for keyword in ["Year Curr", "PHP   0087", "PHP   **** ** ", "                   T                                   T "]):
                            ws[f"O{row_num}"] = cell_value[:4]
                            ws[f"P{row_num}"] = cell_value[5:11]
                            ws[f"Q{row_num}"] = cell_value[11:15]
                            ws[f"R{row_num}"] = cell_value[16:18]
                            ws[f"S{row_num}"] = cell_value[19:21]
                            ws[f"T{row_num}"] = cell_value[21:37]
                            ws[f"U{row_num}"] = cell_value[38:54]
                            ws[f"V{row_num}"] = cell_value[55:57]
                            ws[f"W{row_num}"] = cell_value[57:73]
                            ws[f"X{row_num}"] = cell_value[73:91]
                            ws[f"Y{row_num}"] = cell_value[90:92]
                            ws[f"Z{row_num}"] = cell_value[92:109]
                            ws[f"AA{row_num}"] = cell_value[109:127]
                        elif "Seq.no.  CPU" in cell_value or cell_value[:4] == "0000":
                            ws[f"O{row_num}"] = cell_value[:8]
                            ws[f"P{row_num}"] = cell_value[9:15]
                            ws[f"Q{row_num}"] = cell_value[16:26]
                            ws[f"R{row_num}"] = cell_value[27:33]
                            ws[f"S{row_num}"] = cell_value[34:40]
                            ws[f"T{row_num}"] = cell_value[41:61]
                            ws[f"U{row_num}"] = cell_value[61:101]
                        else:
                            ws[f"P{row_num}"] = cell_value[9:35]
                            ws[f"Q{row_num}"] = cell_value[35:38]
                            ws[f"R{row_num}"] = cell_value[42:43]
                            ws[f"S{row_num}"] = cell_value[44:54]
                            ws[f"T{row_num}"] = cell_value[55:57]
                            ws[f"U{row_num}"] = cell_value[65:75]
                            ws[f"V{row_num}"] = cell_value[60:62]
                            ws[f"W{row_num}"] = cell_value[76:78]
                            ws[f"X{row_num}"] = cell_value[79:95]
                            ws[f"Y{row_num}"] = cell_value[95:99]
                            ws[f"Z{row_num}"] = cell_value[99:115]
                            ws[f"AA{row_num}"] = cell_value[115:130]
    
    
                # Delete columns A to N
                ws.delete_cols(1, 14)
    
                # Apply trim function to all cells
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None and isinstance(cell.value, str):
                            cell.value = cell.value.strip()
    
                # Save the Excel workbook with the same name as the original TXT file
                output_filename = os.path.splitext(filename)[0] + "_GJ converted.xlsx"
                output_path = os.path.join(GJ_output_folder_path, output_filename)
                wb.save(output_path)

        except Exception as e:
            update_status(f"Error {e}.")

# Function to execute the GL conversion and trimming, cannot directly use extract_and_trim(input_path) because of args
def GL_execute_conversion():
    GL_extract_and_trim(GL_input_path_folder)
    status_label.config(text = "Conversion successful.")

# Function to execute the GJ conversion and trimming, cannot directly use extract_and_trim(input_path) because of args
def GJ_execute_conversion():
    GJ_extract_and_trim(GJ_input_folder_path)
    status_label.config(text = "Conversion successful.")


# Create GL execute
execute_button = Button(root, text="GL Convert", command= GL_execute_conversion, padx = 20)
# shoving it into the window
execute_button.pack(side = "left", padx = 10, pady = 10)

# One execute button
aggre_button = Button(root, text = "Convert All", command = lambda:[GL_execute_conversion(), GJ_execute_conversion()], padx = 20)
# shove the button into the window
aggre_button.pack(side = "right", padx = 15, pady = 10)

# Create GJ execute
execute_button2 = Button(root, text = "GJ Convert", command = GJ_execute_conversion, padx = 20)
# shoving it into the window
execute_button2.pack(padx = 10, pady = 10)


# Run the Tkinter event loop
root.mainloop()

 