import os
from tkinter import *
from openpyxl import Workbook

# Specify the path of the files, in this case the Input and Output folder in Conver GL folder
input_path_folder = os.path.join(os.path.expanduser("~"), "OneDrive - Bayer", "Desktop", "File Converter", "GL Input Folder")
output_path_folder = os.path.join(os.path.expanduser("~"), "OneDrive - Bayer", "Desktop", "File Converter", "GL Output Folder")


## GUI script
# Create the window
root = Tk()
root.title("GL Converter")

status_label = Label(root, text="Ready")
# fill = x makes it center
status_label.pack(side = BOTTOM, fill = X, pady = 10)

# function for the status
def update_status(message):
    status_label.config(text = message)

## Main trim function
def extract_and_trim(input_folder):
    for filename in os.listdir(input_folder):
        try:
            if filename.endswith(".txt"):
                # Read content from the TXT file
                with open(os.path.join(input_folder, filename), 'r') as txt_file:
                    content = txt_file.read().splitlines()
    
                # Create a new Excel workbook
                wb = Workbook()
                ws = wb.active
    
                # Apply trimming logic
                for row_num, cell_value in enumerate(content, start=1):
                        if "Document Journal" in cell_value:
                            ws[f"O{row_num}"] = cell_value[:38]
                            ws[f"T{row_num}"] = cell_value[38:81]
                            ws[f"Z{row_num}"] = cell_value[82:112]
                            ws[f"AA{row_num}"] = cell_value[113:134]
                        elif "RFBELJ10" in cell_value:
                            ws[f"O{row_num}"] = cell_value[:96]
                            ws[f"Z{row_num}"] = cell_value[96:116]
                            ws[f"AA{row_num}"] = cell_value[116:]
                        elif "-------------------------------------------------------------------------------------" in cell_value:
                            ws[f"O{row_num}"] = cell_value
                        elif any(keyword in cell_value for keyword in ["Carryfwd", "Pages Total", "Cumulated"]):
                            ws[f"Y{row_num}"] = cell_value[69:83]
                            ws[f"Z{row_num}"] = cell_value[98:114]
                            ws[f"AA{row_num}"] = cell_value[114:130]
                        elif any(keyword in cell_value for keyword in ["Year Curr", "PHP   0087", "PHP   **** ** ", "                   T                                   T "]):
                            ws[f"O{row_num}"] = cell_value[:4]
                            ws[f"P{row_num}"] = cell_value[5:11]
                            ws[f"Q{row_num}"] = cell_value[11:15]
                            ws[f"R{row_num}"] = cell_value[16:18]
                            ws[f"S{row_num}"] = cell_value[19:21]
                            ws[f"T{row_num}"] = cell_value[21:37]
                            ws[f"U{row_num}"] = cell_value[38:54]
                            ws[f"V{row_num}"] = cell_value[55:57]
                            ws[f"W{row_num}"] = cell_value[57:73]
                            ws[f"X{row_num}"] = cell_value[73:91]
                            ws[f"Y{row_num}"] = cell_value[90:92]
                            ws[f"Z{row_num}"] = cell_value[92:109]
                            ws[f"AA{row_num}"] = cell_value[109:127]
                        elif "Seq.no.  CPU" in cell_value or cell_value[:4] == "0000":
                            ws[f"O{row_num}"] = cell_value[:8]
                            ws[f"P{row_num}"] = cell_value[9:15]
                            ws[f"Q{row_num}"] = cell_value[16:26]
                            ws[f"R{row_num}"] = cell_value[27:33]
                            ws[f"S{row_num}"] = cell_value[34:40]
                            ws[f"T{row_num}"] = cell_value[41:61]
                            ws[f"U{row_num}"] = cell_value[61:101]
                        else:
                            ws[f"P{row_num}"] = cell_value[9:35]
                            ws[f"Q{row_num}"] = cell_value[35:38]
                            ws[f"R{row_num}"] = cell_value[42:43]
                            ws[f"S{row_num}"] = cell_value[44:54]
                            ws[f"T{row_num}"] = cell_value[55:57]
                            ws[f"U{row_num}"] = cell_value[65:75]
                            ws[f"V{row_num}"] = cell_value[60:62]
                            ws[f"W{row_num}"] = cell_value[76:78]
                            ws[f"X{row_num}"] = cell_value[79:95]
                            ws[f"Y{row_num}"] = cell_value[95:99]
                            ws[f"Z{row_num}"] = cell_value[99:115]
                            ws[f"AA{row_num}"] = cell_value[115:130]
    
    
                # Delete columns A to N
                ws.delete_cols(1, 14)
    
                # Apply trim function to all cells
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None and isinstance(cell.value, str):
                            cell.value = cell.value.strip()
    
                # Save the Excel workbook with the same name as the original TXT file
                output_filename = os.path.splitext(filename)[0] + "_converted.xlsx"
                output_path = os.path.join(output_path_folder, output_filename)
                wb.save(output_path)

        except Exception as e:
            update_status(f"Error {e}.")

# Function to execute the conversion and trimming, cannot directly use extract_and_trim(input_path) because of args
def execute_conversion():
    extract_and_trim(input_path_folder)
    status_label.config(text = "Conversion successful.")


## Window tool elements
# Create the execute button
execute_button = Button(root, text="Execute", command=execute_conversion, padx = 30)
# shoving it into the window
execute_button.pack(side = "left", padx = 10, pady = 10)

# Create Close button
cancel_button = Button(root, text = "Close", command = root.destroy, padx = 30)
# shoving it into the window
cancel_button.pack(side = "right", padx = 10, pady = 10)

# Run the Tkinter event loop
root.mainloop()

 